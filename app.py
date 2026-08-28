"""
健保資料庫分析系統 (Streamlit 拖曳版 - 廠商分析)
----------------------------------------------
沿用「華典streamlit-0731」的拖曳樞紐分析引擎（streamlit-sortables），
但僅開放「成分、劑型、劑量、規格、廠商」五個欄位供拖曳排列，
且只有一種分析功能（廠商申報量排名），不需要「選擇分析功能」這一步。

步驟：
1. 選擇成分
2. 拖曳欄位到「報表欄位」並排序
3. 逐欄篩選 / 是否合計
4. 選擇要加總的數值欄位（2023~2026年數量），可加占比(%)、成長率(%)
5. 產生 Excel 報表與網頁即時預覽（沿用原本深藍色主題與 A4 橫式列印設定）

需要的套件（requirements.txt）：
    streamlit
    pandas
    openpyxl
    streamlit-sortables
"""

import re
import io
import base64
from datetime import datetime

try:
    from zoneinfo import ZoneInfo
    TAIPEI_TZ = ZoneInfo("Asia/Taipei")
except Exception:
    TAIPEI_TZ = None

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from streamlit_sortables import sort_items
    HAS_SORTABLES = True
except ImportError:
    HAS_SORTABLES = False


# ============================================================
# 基礎設定與資料載入
# ============================================================

st.set_page_config(page_title="健保資料庫分析系統", page_icon="📊", layout="wide")

# ============================================================
# 密碼驗證（進入頁面前）
# ============================================================
# 密碼放在 Streamlit 的 secrets 機制裡，不要寫死在程式碼中：
#   1. 本機開發：在專案根目錄建立 .streamlit/secrets.toml（這個檔案不要上傳到 Git／公開的原始碼），內容例如：
#        APP_PASSWORD = "你的密碼"
#   2. 部署到 Streamlit Community Cloud：到 App 的 Settings → Secrets，貼上同樣的內容即可，
#      不需要另外建立檔案，Cloud 會用同一組 st.secrets 讀取。
# 之後如果要換密碼，只要改 secrets 裡的值，不用改程式碼、也不會留在 Git 歷史紀錄裡。


def check_password() -> bool:
    if st.session_state.get("password_ok", False):
        return True

    st.markdown("## 🔒 請輸入密碼")
    show_pw = st.checkbox("👁️ 顯示密碼", value=False, key="show_pw_toggle")

    with st.form("password_form"):
        pw = st.text_input(
            "密碼", type="default" if show_pw else "password", key="pw_input",
        )
        submitted = st.form_submit_button("登入", type="primary")

    if submitted:
        correct_pw = st.secrets.get("APP_PASSWORD")
        if not correct_pw:
            st.error("⚠️ 尚未設定密碼，請在 .streamlit/secrets.toml (或 Streamlit Cloud 的 Secrets 設定) 加入 APP_PASSWORD。")
        elif pw == correct_pw:
            st.session_state["password_ok"] = True
            st.rerun()
        else:
            st.error("密碼錯誤，請再試一次。")
    return False


if not check_password():
    st.stop()


DATA_FILE = "data2023-2026.csv"

# 只開放這幾個欄位供拖曳排列（依使用者要求，不像 0731 版本開放「所有欄位」）
# 原本的五個維度欄位（成分/劑型/劑量/規格/廠商）之外，新增五個「顯示用」欄位：
# 健保代碼(=藥品代碼)、商品名(=藥品英文名稱)、健保價(=支付價)、ATC碼(=ATC代碼)、
# 許可證連結(=藥品代碼超連結，欄位內容顯示文字固定為「連結」，點擊可開啟原始網頁)
FIXED_FIELDS = ["成分", "劑型", "劑量", "規格", "廠商", "健保代碼", "商品名", "健保價", "ATC碼", "許可證連結"]

# 新欄位的「顯示欄名」-> 「CSV 原始欄名」對照表，載入資料時會把原始欄名改成顯示欄名，
# 這樣後續的拖曳/篩選/分組邏輯可以直接沿用既有五個欄位的處理方式，不用另外寫一套。
EXTRA_FIELD_SOURCE_MAP = {
    "健保代碼": "藥品代號",
    "商品名": "藥品英文名稱",
    "健保價": "支付價",
    "ATC碼": "ATC代碼",
    "許可證連結": "藥品代碼超連結",
}

# 「許可證連結」欄位內容是完整網址，報表上一律只顯示文字「連結」，並把原始網址做成超連結，
# 其餘程式碼用這個常數判斷是否要套用「超連結」的特殊呈現方式。
LINK_FIELD = "許可證連結"
LINK_DISPLAY_TEXT = "連結"

# 原始資料的四個年度數量欄 -> 內部工作用欄名（沿用「申報量」關鍵字，讓占比/成長率引擎能自動辨識）
QTY_COL_MAP = {
    "2023年數量": "2023年申報量",
    "2024年數量": "2024年申報量",
    "2025年數量": "2025年申報量",
    "2026年(1-6月)數量": "2026年申報量",
}
# 內部工作用欄名 -> 顯示用 (年份標籤, 欄位標籤)，2026 特別標示 (1-6月)
QTY_DISPLAY_LABEL = {
    "2023年申報量": ("2023年", "數量"),
    "2024年申報量": ("2024年", "數量"),
    "2025年申報量": ("2025年", "數量"),
    "2026年申報量": ("2026年(1-6月)", "數量"),
}
BLANK_TOKENS = {"nan", "none", "<na>", "null", ""}

# 「含包裹支付醫令量」欄位：CSV 原始欄名 -> 內部工作用欄名。
# 內部欄名刻意保留「申報量」三個字（放在「含包裹」後面），這樣既能沿用既有引擎裡
# 「c 是否含有『申報量』字樣」這類判斷（把它當成一般數值欄一起加總/排序），
# 又能用「含包裹」這個關鍵字把它跟一般（未含包裹）的申報量區分開來，
# 讓占比/成長率各自獨立計算、不會互相混用分母。
BUNDLE_QTY_COL_MAP = {
    "2023年數量(含包裹)": "2023年含包裹申報量",
    "2024年數量(含包裹)": "2024年含包裹申報量",
    "2025年數量(含包裹)": "2025年含包裹申報量",
    "2026年(1-6月)數量(含包裹)": "2026年含包裹申報量",
}
QTY_COL_MAP.update(BUNDLE_QTY_COL_MAP)
QTY_DISPLAY_LABEL.update({
    "2023年含包裹申報量": ("2023年", "數量(含包裹)"),
    "2024年含包裹申報量": ("2024年", "數量(含包裹)"),
    "2025年含包裹申報量": ("2025年", "數量(含包裹)"),
    "2026年含包裹申報量": ("2026年(1-6月)", "數量(含包裹)"),
})

# 健保署官方定義，顯示在第4步供使用者參考「數量」與「數量(含包裹)」的差別。
QTY_DEFINITION_NOTE = (
    "ℹ️ 「醫令量」係加總醫療院所申報之計價醫令；"
    "「含包裹支付醫令量」係加總醫療院所申報前述計價醫令，"
    "及其他不計價醫令（如：以日劑藥費 MA1~MA4 醫令申報費用，併同申報所開立藥品代碼及其醫令量）。"
)

# 2026 年推估全年數量（以 1-5 月數據 ÷5×12 估算），以及對應的推估占比／推估成長率欄位。
# 這三個欄位不是原始 CSV 資料，是從「2026年申報量」衍生出來的，所以獨立用常數命名，
# 不放進 QTY_COL_MAP（那是給「讀取 CSV 原始欄位」用的），避免使用者在第4步選數值欄位時誤選到它。
EST_QTY_COL = "2026年推估申報量"
EST_PCT_COL = "2026推估占比(%)"
EST_GROWTH_COL = "2025-2026推估成長率(%)"
# 讓「推估占比／推估成長率」可以像其他年度一樣，直接出現在「加入年度占比」「加入年度成長率」
# 的選單裡讓使用者勾選（而不是只要勾了推估數量就自動夾帶），這兩個是選單裡專用的識別值
EST_PCT_YEAR = "2026推估"
EST_GROWTH_PAIR = ("2025", "2026推估")
# 讓「2026年推估數量」可以跟其他年度數量一樣，直接出現在第4步的「選擇要加總的數值欄位」
# 選單裡（預設勾選），不用再另外一個獨立的勾選框
QTY_DISPLAY_LABEL[EST_QTY_COL] = ("2026年推估", "數量")

# 「2026年含包裹推估數量」：同樣以 1-6 月「含包裹」數量 ÷6×12 估算，是 EST_QTY_COL 的「含包裹」版本，
# 邏輯與命名方式完全比照上面的一般推估數量，只是全部多套一層「含包裹」。
EST_BUNDLE_QTY_COL = "2026年含包裹推估申報量"
EST_BUNDLE_PCT_COL = "2026含包裹推估占比(%)"
EST_BUNDLE_GROWTH_COL = "2025-2026含包裹推估成長率(%)"
EST_BUNDLE_PCT_YEAR = "2026含包裹推估"
# 成長率的起點也要標成「含包裹」，確保「含包裹推估成長率」永遠是「含包裹2025 → 含包裹2026推估」
# 這種同類型的比較，不會出現「不含包裹2025 → 含包裹2026推估」這種跨類型、容易誤導的組合。
EST_BUNDLE_GROWTH_PAIR = ("2025含包裹", "2026含包裹推估")
QTY_DISPLAY_LABEL[EST_BUNDLE_QTY_COL] = ("2026年推估", "數量(含包裹)")

# 年度占比／年度成長率的「年度代碼」：真實年度一律用純西元年字串（例如 "2023"）代表醫令量，
# 加上「含包裹」三個字代表含包裹醫令量（例如 "2023含包裹"）；「2026推估」「2026含包裹推估」則是
# 兩個推估用的特殊代碼常數（就是上面的 EST_PCT_YEAR / EST_BUNDLE_PCT_YEAR）。
# 這一組代碼統一給「加入年度占比」「加入年度成長率」的合併選單使用，讓使用者能在同一個下拉選單
# 裡看到「2025」「2025含包裹」「2026推估」「2026含包裹推估」等選項，不用分開兩組選單選。
QTY_YEARS = ["2023", "2024", "2025", "2026"]
BUNDLE_SUFFIX = "含包裹"


def is_bundle_code(code: str) -> bool:
    return code in (EST_BUNDLE_PCT_YEAR,) or (code not in (EST_PCT_YEAR,) and code.endswith(BUNDLE_SUFFIX))


def code_to_qty_col(code: str) -> str:
    """把年度代碼（例如 "2025"、"2025含包裹"、"2026推估"、"2026含包裹推估"）轉成對應的內部數量欄名。"""
    if code == EST_PCT_YEAR:
        return EST_QTY_COL
    if code == EST_BUNDLE_PCT_YEAR:
        return EST_BUNDLE_QTY_COL
    if code.endswith(BUNDLE_SUFFIX):
        y = code[: -len(BUNDLE_SUFFIX)]
        return f"{y}年含包裹申報量"
    return f"{code}年申報量"


def code_to_pct_col(code: str) -> str:
    """把年度代碼轉成對應的占比欄名。"""
    if code == EST_PCT_YEAR:
        return EST_PCT_COL
    if code == EST_BUNDLE_PCT_YEAR:
        return EST_BUNDLE_PCT_COL
    if code.endswith(BUNDLE_SUFFIX):
        y = code[: -len(BUNDLE_SUFFIX)]
        return f"{y}年含包裹占比(%)"
    return f"{code}年占比(%)"


def codes_to_growth_col(y1_code: str, y2_code: str) -> str:
    """把一組年度代碼區間 (y1, y2) 轉成對應的成長率欄名。y1、y2 一定同屬「醫令量」或「含包裹醫令量」。"""
    if y2_code == EST_PCT_YEAR:
        return EST_GROWTH_COL
    if y2_code == EST_BUNDLE_PCT_YEAR:
        return EST_BUNDLE_GROWTH_COL
    is_pkg = y1_code.endswith(BUNDLE_SUFFIX)
    yy1 = y1_code[: -len(BUNDLE_SUFFIX)] if is_pkg else y1_code
    yy2 = y2_code[: -len(BUNDLE_SUFFIX)] if is_pkg else y2_code
    return f"{yy1}-{yy2}年含包裹成長率(%)" if is_pkg else f"{yy1}-{yy2}年成長率(%)"


# 合併後的「加入年度占比」選單選項：依序放「2023」「2023含包裹」「2024」「2024含包裹」……，
# 最後才是兩個推估選項，跟使用者是否有把對應的數量欄位拖曳到「要加總的數值欄位」完全無關——
# 只要選了年度，該年度（該類型）的占比／成長率就會用該年度（該類型）的原始資料正確算出來，
# 不會因為使用者沒有另外勾選顯示那一欄數量，就讓分母抓不到資料而變成 0。
PCT_YEAR_CODE_OPTIONS = []
for _y in QTY_YEARS:
    PCT_YEAR_CODE_OPTIONS.append(_y)
    PCT_YEAR_CODE_OPTIONS.append(_y + BUNDLE_SUFFIX)
PCT_YEAR_CODE_OPTIONS += [EST_PCT_YEAR, EST_BUNDLE_PCT_YEAR]

# 合併後的「加入年度成長率」選單選項：只有「醫令量→醫令量」或「含包裹醫令量→含包裹醫令量」
# 這種同類型的年度區間，不會出現「醫令量→含包裹醫令量」這種跨類型比較（會誤導使用者）。
# 2026 年只有 1-6 月的資料，實際成長率不提供「XX年→2026年(1-6月)」這種全年比半年的組合，
# 只保留「XX年→2026年推估」／「XX含包裹→2026含包裹推估」。
GROWTH_PAIR_CODE_OPTIONS = (
    [(QTY_YEARS[i], QTY_YEARS[i + 1]) for i in range(len(QTY_YEARS) - 1) if QTY_YEARS[i + 1] != "2026"]
    + [EST_GROWTH_PAIR]
    + [
        (QTY_YEARS[i] + BUNDLE_SUFFIX, QTY_YEARS[i + 1] + BUNDLE_SUFFIX)
        for i in range(len(QTY_YEARS) - 1) if QTY_YEARS[i + 1] != "2026"
    ]
    + [EST_BUNDLE_GROWTH_PAIR]
)


def pct_year_code_label(code: str) -> str:
    """年度占比選單的顯示文字，例如 "2025" -> "2025年"、"2025含包裹" -> "2025年含包裹"、
    "2026" -> "2026年(1-6月)"（只有半年資料，特別標示避免誤認為全年占比）。"""
    if code == EST_PCT_YEAR:
        return "2026年推估"
    if code == EST_BUNDLE_PCT_YEAR:
        return "2026年含包裹推估"
    if code.endswith(BUNDLE_SUFFIX):
        y = code[: -len(BUNDLE_SUFFIX)]
        return f"{y}年(1-6月)含包裹" if y == "2026" else f"{y}年含包裹"
    return f"{code}年(1-6月)" if code == "2026" else f"{code}年"


def growth_pair_code_label(y1_code: str, y2_code: str) -> str:
    """年度成長率選單的顯示文字，例如 ("2023","2024") -> "2023年 → 2024年"、
    ("2025含包裹","2026含包裹推估") -> "2025年含包裹 → 2026年含包裹推估"。"""
    def fmt(code):
        if code == EST_PCT_YEAR:
            return "2026年推估"
        if code == EST_BUNDLE_PCT_YEAR:
            return "2026年含包裹推估"
        if code.endswith(BUNDLE_SUFFIX):
            y = code[: -len(BUNDLE_SUFFIX)]
            return f"{y}年(1-6月)含包裹" if y == "2026" else f"{y}年含包裹"
        return f"{code}年(1-6月)" if code == "2026" else f"{code}年"

    return f"{fmt(y1_code)} → {fmt(y2_code)}"


@st.cache_data(show_spinner=False)
def load_data(path: str) -> pd.DataFrame:
    try:
        df = pd.read_csv(path)
    except Exception as e:
        st.error(f"⚠️ 讀取資料檔案「{path}」時發生錯誤：{e}")
        return pd.DataFrame()

    df.columns = df.columns.str.strip()

    # 把新增欄位的原始 CSV 欄名改成報表要用的顯示欄名（例如「藥品代號」→「健保代碼」），
    # 之後 FIXED_FIELDS 相關的邏輯就能跟原本五個欄位一樣統一處理。
    rename_map = {src: disp for disp, src in EXTRA_FIELD_SOURCE_MAP.items() if src in df.columns}
    df = df.rename(columns=rename_map)

    # 清洗所有可拖曳欄位：去除空白，並把各種「無資料」寫法統一成空字串
    # （「許可證連結」欄位內容是網址，保留原始網址字串，不做 lower() 比對以外的處理）
    for col in FIXED_FIELDS:
        if col not in df.columns:
            df[col] = ""
            continue
        cleaned = df[col].astype(str).str.strip()
        if col == "劑量":
            # 移除劑量字串內所有空白（含中間空格），避免 "250mg" 與 "250 mg" 被當成不同劑量
            cleaned = cleaned.str.replace(r"\s+", "", regex=True)
        cleaned = cleaned.apply(lambda x: "" if str(x).lower() in BLANK_TOKENS else x)
        df[col] = cleaned

    # 四個年度數量欄：清除千分位逗號/空白後轉數字，轉出內部工作欄名
    for raw_col, internal_col in QTY_COL_MAP.items():
        if raw_col not in df.columns:
            df[raw_col] = 0
        cleaned = df[raw_col].astype(str).str.replace(",", "", regex=False).str.strip()
        df[internal_col] = pd.to_numeric(cleaned, errors="coerce").fillna(0)

    # 2026 年推估全年數量：用 1-5 月數量 ÷5×12 估算。先在這裡逐列算好，
    # 之後不管怎麼篩選/分組加總，用一般的加總邏輯就能得到正確的推估總量
    # （因為「先加總再乘」跟「先乘再加總」對於固定倍率而言結果相同）。
    df[EST_QTY_COL] = df["2026年申報量"] * 12 / 6
    df[EST_BUNDLE_QTY_COL] = df["2026年含包裹申報量"] * 12 / 6

    return df


def parse_dosage(d):
    m = re.search(r"[\d.]+", str(d))
    return float(m.group()) if m else 0.0


def order_display_cols(value_cols, pct_cols, growth_cols):
    """占比要接在數量後面，而不是放在所有數值欄位最後"""
    qty_cols = [c for c in value_cols if "申報量" in c]
    other_cols = [c for c in value_cols if c not in qty_cols]
    return qty_cols + pct_cols + other_cols + growth_cols


def pretty_header(col: str):
    """把內部欄名轉成「多行標題」的字串清單（2或3行），呼叫端自行用 <br>（HTML）或 \\n（Excel）組合。
    2026年(1-6月)的數量/占比改成三行顯示；占比/成長率欄位的年份／年度區間都保留「年」字。"""
    if col == "2026年申報量":
        return ["2026年", "(1-6月)", "數量"]
    if col == "2026年含包裹申報量":
        return ["2026年", "(1-6月)", "數量(含包裹)"]
    if col == EST_QTY_COL:
        return ["2026年", "推估數量"]
    if col in QTY_DISPLAY_LABEL:
        return list(QTY_DISPLAY_LABEL[col])

    if col == "2026年占比(%)":
        return ["2026年", "(1-6月)", "占比(%)"]
    if col == "2026年含包裹占比(%)":
        return ["2026年", "(含包裹)(1-6月)", "占比(%)"]
    if col == EST_PCT_COL:
        return ["2026年", "推估占比(%)"]
    if col == EST_BUNDLE_PCT_COL:
        return ["2026年", "(含包裹)", "推估占比(%)"]
    m = re.match(r"^(\d{4})年含包裹占比\(%\)$", col)
    if m:
        y = m.group(1)
        return [f"{y}年", "(含包裹)", "占比(%)"]
    m = re.match(r"^(\d{4})年占比\(%\)$", col)
    if m:
        y = m.group(1)
        return [f"{y}年", "占比(%)"]

    if col == EST_GROWTH_COL:
        return ["2025-2026年", "推估成長率(%)"]
    if col == EST_BUNDLE_GROWTH_COL:
        return ["2025-2026年", "(含包裹)", "推估成長率(%)"]
    m = re.match(r"^(\d{4})-(\d{4})年含包裹成長率\(%\)$", col)
    if m:
        y1, y2 = m.group(1), m.group(2)
        y1d = f"{y1}(1-6月)" if y1 == "2026" else y1
        y2d = f"{y2}(1-6月)" if y2 == "2026" else y2
        return [f"{y1d}-{y2d}年", "(含包裹)", "成長率(%)"]
    m = re.match(r"^(\d{4})-(\d{4})年成長率\(%\)$", col)
    if m:
        y1, y2 = m.group(1), m.group(2)
        y1d = f"{y1}(1-6月)" if y1 == "2026" else y1
        y2d = f"{y2}(1-6月)" if y2 == "2026" else y2
        return [f"{y1d}-{y2d}年", "成長率(%)"]

    return [col]


# ============================================================
# 通用樞紐分析引擎：任意欄位順序 + 任意層級巢狀合計 + 占比/成長率
# （沿用 0731 版本的引擎邏輯，因資料本身已是乾淨數字，移除原本的
#   「無資料(-)/異常標記文字」保留顯示機制，只保留核心的巢狀分組/合計）
# ============================================================

def build_nested_rows(df: pd.DataFrame, row_fields: list, subtotal_fields: list, value_cols: list,
                       pct_year_codes: list = None, growth_pair_codes: list = None):
    pct_year_codes = pct_year_codes or []  # 合併後的年度占比代碼，例如 ["2025", "2025含包裹", "2026推估"]
    growth_pair_codes = growth_pair_codes or []  # 合併後的成長率年度區間代碼，例如 [("2023","2024"), ("2025含包裹","2026含包裹推估")]

    # value_cols 是使用者勾選「要顯示成欄位」的數量欄。占比／成長率所需要的分母／分子欄位，
    # 不管使用者有沒有勾選要「顯示」該欄數量，只要選了對應的年度占比／成長率，就要一併算進來
    # （這樣才不會發生「使用者只顯示了2026年含包裹數量，卻要看2025含包裹占比」時，
    #  因為2025年含包裹申報量沒被加總、分母變成0，導致占比/成長率被誤算成0%的情況）。
    display_value_cols = list(value_cols)
    calc_cols = list(display_value_cols)

    def _ensure_calc_col(c):
        if c not in calc_cols:
            calc_cols.append(c)

    for code in pct_year_codes:
        _ensure_calc_col(code_to_qty_col(code))
    for (y1_code, y2_code) in growth_pair_codes:
        _ensure_calc_col(code_to_qty_col(y1_code))
        _ensure_calc_col(code_to_qty_col(y2_code))

    # 先依「報表欄位」的組合彙總數值，確保同一個欄位組合只會出現一列
    df = df.groupby(row_fields, as_index=False)[calc_cols].sum()

    # 依序把使用者選的年度占比／成長率代碼轉成實際的欄名，並各自記錄「要拿哪個數量欄當分母／分子」，
    # 之後計算階段直接查表，不必再從欄名字串反推年度／醫令量或含包裹醫令量。
    pct_cols, pct_base_map = [], {}
    for code in pct_year_codes:
        col_name = code_to_pct_col(code)
        pct_cols.append(col_name)
        pct_base_map[col_name] = code_to_qty_col(code)

    growth_cols, growth_base_map = [], {}
    for (y1_code, y2_code) in growth_pair_codes:
        col_name = codes_to_growth_col(y1_code, y2_code)
        growth_cols.append(col_name)
        growth_base_map[col_name] = (code_to_qty_col(y1_code), code_to_qty_col(y2_code))

    def compute_extra_for_row(sums, top_totals):
        extra = {}
        for c in pct_cols:
            base = pct_base_map[c]
            extra[c] = sums.get(base, 0) / top_totals.get(base, 0) if top_totals.get(base, 0) else 0
        for c in growth_cols:
            b1, b2 = growth_base_map[c]
            extra[c] = (sums.get(b2, 0) - sums.get(b1, 0)) / sums.get(b1, 0) if sums.get(b1, 0) else 0
        return extra

    def compute_extra_for_group(totals):
        extra = {c: 1.0 for c in pct_cols}  # 小計/總計列本身佔比恆為 100%
        for c in growth_cols:
            b1, b2 = growth_base_map[c]
            extra[c] = (totals.get(b2, 0) - totals.get(b1, 0)) / totals.get(b1, 0) if totals.get(b1, 0) else 0
        return extra

    # 排序用的「數量欄」跟上面判斷年度用的 qty_cols 不同：排序要把「2026年推估數量」也算進去，
    # 這樣即使使用者只選了推估數量、沒選任何實際年度欄位，排名（由大到小）仍然正常運作
    sort_qty_cols = [c for c in calc_cols if "申報量" in c]
    year_cols = sorted(sort_qty_cols, reverse=True)

    sort_cols, sort_asc, temp_sort_cols = [], [], []
    for col in subtotal_fields:
        if col == "劑量":
            # 劑量欄需依實際數值排序 (例如 12mg/ml < 24mg/ml < 100mg/ml)，而非字母排序
            tmp_col = "__sortkey_劑量"
            df[tmp_col] = df[col].map(parse_dosage)
            sort_cols.append(tmp_col)
            temp_sort_cols.append(tmp_col)
        elif col == "規格":
            # 規格欄同樣需依數值排序 (例如 50ml < 100ml < 150ml)，而非字母排序，
            # 否則字串比較會把 "100ml"、"150ml" 排在 "50ml" 之前（因為開頭字元 "1" < "5"）
            tmp_col = "__sortkey_規格"
            df[tmp_col] = df[col].map(parse_dosage)
            sort_cols.append(tmp_col)
            temp_sort_cols.append(tmp_col)
        else:
            sort_cols.append(col)
        sort_asc.append(True)
    sort_cols += year_cols
    sort_asc += [False] * len(year_cols)

    df_sorted = df.sort_values(by=sort_cols, ascending=sort_asc) if sort_cols else df
    if temp_sort_cols:
        df_sorted = df_sorted.drop(columns=temp_sort_cols)

    # 「省略重複值」規則（類似 Excel 樞紐分析表的大綱模式）：
    # 由左到右依序比較每一欄，只要目前這一欄的值與上一列相同、且左邊所有欄位也都跟上一列相同
    # （代表仍在同一個群組內），就把這一欄留白；只要有一欄改變了，從那一欄開始（含）到最右邊都要照常顯示，
    # 不能因為剛好文字相同就被誤判成同一群組而被省略。
    # 這一段取代舊版只針對「主要分組欄位」或「全表僅有一種值」才省略重複的作法，
    # 讓「劑型」「規格」這類欄位在同一組內重複出現時，也能正確被省略，不會重複印出。
    _MISSING = object()

    rows = []
    last_vals = {}
    first_flags = {c: True for c in subtotal_fields}

    def emit_row(row, top_totals):
        rec_vals = {}
        prefix_broken = False
        for f in row_fields:
            val = row[f]
            if f in subtotal_fields:
                if first_flags[f]:
                    show = val
                    prefix_broken = True  # 新的合計群組開始，之後的欄位一律照常顯示
                else:
                    show = ""
            else:
                if (not prefix_broken) and val == last_vals.get(f, _MISSING):
                    show = ""
                else:
                    show = val
                    prefix_broken = True
            last_vals[f] = val
            rec_vals[f] = show
        # calc_sums 用來算占比/成長率（含隱藏未顯示的 EST_QTY_COL），
        # 但實際要顯示出來的欄位只有 display_value_cols（使用者真正勾選要顯示的數量欄）
        calc_sums = {c: row[c] for c in calc_cols}
        sums = {c: row[c] for c in display_value_cols}
        sums.update(compute_extra_for_row(calc_sums, top_totals))
        rows.append({"type": "data", "values": rec_vals, "sums": sums})
        for c in first_flags:
            first_flags[c] = False

    def emit_subtotal(field, label, totals):
        rec_vals = {f: "" for f in row_fields}
        rec_vals[field] = f"{label} 合計"
        full = {c: totals.get(c, 0) for c in display_value_cols}
        full.update(compute_extra_for_group(totals))
        rows.append({"type": "subtotal", "values": rec_vals, "sums": full})

    def recurse(sub_df, level_idx, top_totals):
        if level_idx >= len(subtotal_fields):
            for _, row in sub_df.iterrows():
                emit_row(row, top_totals)
            return
        col = subtotal_fields[level_idx]
        for _, grp in sub_df.groupby(col, sort=False):
            totals = {c: grp[c].sum() for c in calc_cols}
            next_top = totals if level_idx == len(subtotal_fields) - 1 else top_totals
            label = grp[col].iloc[0]
            first_flags[col] = True
            recurse(grp, level_idx + 1, next_top)
            emit_subtotal(col, label, totals)

    grand_totals = {c: df[c].sum() for c in calc_cols}
    if subtotal_fields:
        recurse(df_sorted, 0, grand_totals)
    else:
        for _, row in df_sorted.iterrows():
            emit_row(row, grand_totals)

    total_vals = {f: "" for f in row_fields}
    if row_fields:
        total_vals[row_fields[0]] = "總計"
    full_grand = {c: grand_totals.get(c, 0) for c in display_value_cols}
    full_grand.update(compute_extra_for_group(grand_totals))
    rows.append({"type": "total", "values": total_vals, "sums": full_grand})

    return rows, pct_cols, growth_cols


def get_report_timestamp() -> str:
    now = datetime.now(TAIPEI_TZ) if TAIPEI_TZ else datetime.now()
    return f"{now.year}/{now.month}/{now.day} {now.hour:02d}:{now.minute:02d}"


def safe_numeric(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            if pd.isna(v):
                return 0.0, True
        except (TypeError, ValueError):
            pass
        return float(v), True
    try:
        n = pd.to_numeric(v)
        return float(n), True
    except (TypeError, ValueError):
        return v, False


# ============================================================
# 樣式化 HTML 預覽（沿用原 Gradio 版本的深藍色主題：#1F497D 標題、
# #DCE6F1 小計、#B8CCE4 總計）
# ============================================================

# 字型統一使用同一組字型堆疊。這裡刻意把 Google Fonts 載入的「Noto Sans TC」放在最前面，
# 而不是微軟正黑體/Microsoft JhengHei：
#   之前雖然三個字型都放進了同一個堆疊，但微軟正黑體/Microsoft JhengHei 通常沒有「真正設計的
#   粗體字檔」，瀏覽器/系統只能用「模擬粗體(faux bold)」硬把筆畫加粗，這種模擬粗體的字形
#   （尤其是英文字母）跟原本的細體看起來會明顯不一樣，才會出現「標題(粗體)」跟「表格內容(細體)」
#   字型風格不一致的狀況。改成優先用 Noto Sans TC 之後，因為 Google Fonts 有把 400(細體)跟
#   700(粗體) 這兩個字重各自獨立設計、獨立下載，粗體不是用模擬的，所以標題、表頭、數字、頁尾
#   不管字重粗細，看起來都會是同一套字型、同一種風格。
FONT_FAMILY = "'Noto Sans TC', '微軟正黑體', 'Microsoft JhengHei', sans-serif"
HEADER_COLOR = "#1F497D"
SUBTOTAL_COLOR = "#DCE6F1"
TOTAL_COLOR = "#B8CCE4"


def build_html_table(rows, row_fields, value_cols, pct_cols, growth_cols, report_title):
    extra_cols = set(pct_cols + growth_cols)
    display_cols = order_display_cols(value_cols, pct_cols, growth_cols)
    headers = row_fields + display_cols

    th_style = (
        f"color:#FFFFFF;background-color:{HEADER_COLOR};text-align:center;"
        f"border:1px solid #D9D9D9;padding:10px 12px;font-weight:bold;font-family:{FONT_FAMILY};"
    )
    td_style = f"border:1px solid #D9D9D9;padding:8px 12px;font-family:{FONT_FAMILY};white-space:nowrap;"

    # 同類型欄位（數量／占比／成長率）欄寬要一致，用 min-width 統一設定
    def col_min_width(c):
        if c in growth_cols:
            return "110px"
        if c in pct_cols:
            return "95px"
        return "115px"  # 數量欄

    html = [
        "<div class='report-preview-wrapper'><div class='report-container'>",
        f"<h2 class='report-title'>{report_title}</h2>",
        "<div class='table-responsive'>",
        f"<table style='border-collapse:collapse;width:max-content;min-width:100%;font-size:15px;font-family:{FONT_FAMILY};'>",
        "<tr>",
    ]
    for f in row_fields:
        html.append(f"<th style='{th_style}'>{f}</th>")
    for c in display_cols:
        lines = pretty_header(c)
        html.append(f"<th style='{th_style}min-width:{col_min_width(c)};'>{'<br>'.join(lines)}</th>")
    html.append("</tr>")

    for r in rows:
        bg = ""
        fw = "font-weight:normal;"
        if r["type"] == "subtotal":
            bg = f"background-color:{SUBTOTAL_COLOR};"
            fw = "font-weight:bold;"
        elif r["type"] == "total":
            bg = f"background-color:{TOTAL_COLOR};"
            fw = "font-weight:bold;"
        html.append(f"<tr style='{bg}{fw}'>")
        for i, f in enumerate(row_fields):
            v = r["values"].get(f, "")
            align = "left" if i == 0 else "center"
            if f == LINK_FIELD and v not in ("", None):
                cell_html = f"<a href='{v}' target='_blank' rel='noopener' style='color:#1F497D;'>{LINK_DISPLAY_TEXT}</a>"
            else:
                cell_html = v
            html.append(f"<td style='{td_style}text-align:{align};'>{cell_html}</td>")
        for c in display_cols:
            if c in extra_cols:
                v, ok = safe_numeric(r["sums"].get(c, 0))
                html.append(f"<td style='{td_style}text-align:right;'>{v:.1%}</td>" if ok else f"<td style='{td_style}text-align:right;'>{v}</td>")
            else:
                raw = r["sums"].get(c, "")
                if raw == "":
                    html.append(f"<td style='{td_style}'></td>")
                else:
                    v, ok = safe_numeric(raw)
                    html.append(f"<td style='{td_style}text-align:right;'>{v:,.0f}</td>" if ok else f"<td style='{td_style}text-align:right;'>{v}</td>")
        html.append("</tr>")

    html.append("</table></div>")
    html.append(
        "<div class='report-footer'>"
        "<span>中央健康保險署  政府資料開放平台 2026年資料</span>"
        "<span class='report-footer-links'>"
        "<span>https://data.gov.tw/dataset/22131</span>"
        "<span>https://data.gov.tw/dataset/23715</span>"
        "</span>"
        "</div>"
    )
    html.append("</div></div>")

    # 網頁即時預覽（st.markdown）跟匯出 PNG（components.html 的 iframe）是兩個不同的渲染環境，
    # 之前 Google Fonts 的 <link> 只加在匯出 PNG 用的 CAPTURE_HTML_TEMPLATE 裡，網頁預覽本身
    # 沒有載入 Noto Sans TC，這裡一併加上，確保「畫面上看到的預覽」跟「匯出的 PNG」字型一致。
    font_link = (
        '<link rel="preconnect" href="https://fonts.googleapis.com">'
        '<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>'
        '<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@400;700&display=swap" rel="stylesheet">'
    )

    style_block = font_link + f"""
    <style>
        .report-preview-wrapper {{ padding: 15px 0; }}
        .report-container {{
            background-color: #FFFFFF !important;
            color: #333333 !important;
            padding: 20px !important;
            width: 100%;
            box-sizing: border-box;
            border-radius: 12px;
            box-shadow: 0 8px 24px rgba(0,0,0,0.15);
            border: 1px solid #E0E0E0;
        }}
        .table-responsive {{ width: 100%; overflow-x: auto; -webkit-overflow-scrolling: touch; padding-bottom: 10px; }}
        .report-title {{
            color: #000000 !important; text-align: center; font-family: {FONT_FAMILY};
            font-weight: bold; margin-top: 5px; margin-bottom: 20px;
            white-space: normal; word-break: break-word; overflow-wrap: break-word;
            /* 讓標題不要撐大外層容器的寬度：容器寬度由表格(max-content)決定，
               標題本身用 width:0; min-width:100% 的技巧強制在容器現有寬度內換行 */
            display: block; width: 0; min-width: 100%; box-sizing: border-box;
        }}
        .report-footer {{
            display: flex; flex-wrap: wrap; justify-content: space-between;
            margin-top: 25px; padding-top: 15px; border-top: 1px solid #D9D9D9;
            font-size: 13px !important; font-family: {FONT_FAMILY}; font-weight: bold; color: #333333 !important;
        }}
        .report-footer-links {{
            display: flex; flex-direction: column; align-items: flex-end; text-align: right;
        }}
    </style>
    """
    return style_block + "".join(html)


CAPTURE_HTML_TEMPLATE = """
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@400;700&display=swap" rel="stylesheet">
<script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
<div style="text-align:center;">
  <button id="export-png-btn-{uid}" style="background-color:#1F497D;color:white;border:none;padding:10px 20px;
    border-radius:6px;font-size:14px;cursor:pointer;width:100%;">🖼️ 匯出 PNG 圖片</button>
</div>
<div id="result-area-{uid}" style="margin-top:10px;font-size:13px;color:#1F497D;text-align:center;"></div>
<div id="capture-wrap-{uid}" style="position:absolute; left:-99999px; top:0; width:max-content;">{table_html}</div>
<script>
(function() {{
    const resultArea = document.getElementById('result-area-{uid}');
    const pngBtn = document.getElementById('export-png-btn-{uid}');
    let busy = false;

    async function captureCanvas() {{
        const target = document.getElementById('capture-wrap-{uid}');
        if (document.fonts && document.fonts.ready) {{ await document.fonts.ready; }}
        if (typeof html2canvas === 'undefined') {{ throw new Error('html2canvas 尚未載入完成，請確認網路連線後再試一次'); }}
        return await html2canvas(target, {{
            scale: 2, backgroundColor: '#ffffff',
            windowWidth: target.scrollWidth, windowHeight: target.scrollHeight,
            width: target.scrollWidth, height: target.scrollHeight, useCORS: true
        }});
    }}

    async function shareOrDownload(blob, filename, mime) {{
        const file = new File([blob], filename, {{type: mime}});
        const isMobile = /iPhone|iPad|iPod|Android/i.test(navigator.userAgent) ||
            (navigator.platform === 'MacIntel' && navigator.maxTouchPoints > 1);
        if (isMobile) {{
            if (navigator.canShare && navigator.canShare({{files: [file]}})) {{
                try {{ await navigator.share({{files: [file]}}); resultArea.innerText = '✅ 已開啟分享選單'; }}
                catch (shareErr) {{ }}
            }} else {{
                const url = URL.createObjectURL(blob);
                resultArea.innerHTML = "目前瀏覽器不支援原生分享，請長按下方圖片選擇「儲存影像」：<br/>" +
                    "<img src='" + url + "' style='max-width:100%;border-radius:8px;border:1px solid #eee;margin-top:6px;' />";
            }}
        }} else {{
            const url = URL.createObjectURL(blob);
            const a = document.createElement('a');
            a.href = url; a.download = filename;
            document.body.appendChild(a); a.click(); document.body.removeChild(a);
            setTimeout(function() {{ URL.revokeObjectURL(url); }}, 5000);
            resultArea.innerText = '✅ 已開始下載';
        }}
    }}

    pngBtn.addEventListener('click', async function() {{
        if (busy) return;
        busy = true; pngBtn.disabled = true;
        resultArea.innerText = '🔄 產生圖片中，請稍候...';
        try {{
            const canvas = await captureCanvas();
            const blob = await new Promise(function(resolve) {{ canvas.toBlob(resolve, 'image/png'); }});
            if (!blob) {{ throw new Error('圖片產生失敗 (canvas 轉換為空)'); }}
            await shareOrDownload(blob, '{filename}.png', 'image/png');
        }} catch (err) {{
            resultArea.innerText = '❌ 匯出失敗：' + (err && err.message ? err.message : err);
        }} finally {{ busy = false; pngBtn.disabled = false; }}
    }});
}})();
</script>
"""

EXCEL_SHARE_HTML_TEMPLATE = """
<div style="text-align:center;">
  <button id="excel-share-btn-{uid}" style="background-color:#1F497D;color:white;border:none;padding:10px 20px;
    border-radius:6px;font-size:14px;cursor:pointer;width:100%;">📄 下載 / 分享 Excel 報表</button>
</div>
<script>
(function() {{
    const btn = document.getElementById('excel-share-btn-{uid}');
    let busy = false;
    btn.addEventListener('click', async function() {{
        if (busy) return;
        busy = true; btn.disabled = true;
        try {{
            const b64 = "{b64data}";
            const byteChars = atob(b64);
            const byteNumbers = new Array(byteChars.length);
            for (let i = 0; i < byteChars.length; i++) {{ byteNumbers[i] = byteChars.charCodeAt(i); }}
            const byteArray = new Uint8Array(byteNumbers);
            const mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet';
            const blob = new Blob([byteArray], {{type: mime}});
            const file = new File([blob], '{filename}.xlsx', {{type: mime}});
            const isMobile = /iPhone|iPad|iPod|Android/i.test(navigator.userAgent) ||
                (navigator.platform === 'MacIntel' && navigator.maxTouchPoints > 1);
            if (isMobile && navigator.canShare && navigator.canShare({{files: [file]}})) {{
                try {{ await navigator.share({{files: [file]}}); }} catch (shareErr) {{ }}
            }} else {{
                const url = URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url; a.download = '{filename}.xlsx';
                document.body.appendChild(a); a.click(); document.body.removeChild(a);
                setTimeout(function() {{ URL.revokeObjectURL(url); }}, 5000);
            }}
        }} finally {{ busy = false; btn.disabled = false; }}
    }});
}})();
</script>
"""


def render_excel_share(excel_bytes: bytes, filename: str, uid: str):
    b64data = base64.b64encode(excel_bytes).decode()
    safe_uid = re.sub(r"[^0-9A-Za-z_]", "_", uid)
    components.html(EXCEL_SHARE_HTML_TEMPLATE.format(b64data=b64data, filename=filename, uid=safe_uid), height=50)


# ============================================================
# Excel 匯出（沿用原 Gradio 版本的深藍色主題、A4 橫式、列印設定）
# ============================================================

def generate_excel_bytes(rows, row_fields, value_cols, pct_cols, growth_cols, report_title):
    extra_cols = set(pct_cols + growth_cols)
    display_cols = order_display_cols(value_cols, pct_cols, growth_cols)
    headers = row_fields + display_cols

    wb = Workbook()
    ws = wb.active
    ws.title = "廠商分析"

    ws.views.sheetView[0].showGridLines = False
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = "landscape"
    ws.print_options.horizontalCentered = True
    ws.page_setup.scaleWithDoc = True
    ws.page_setup.alignWithMargins = True
    # 沒有這一行，Excel 會忽略 fitToWidth/fitToHeight，直接照 100% 實際大小分頁，
    # 導致欄位很多時被切成好幾頁 —— 必須明確開啟 fitToPage 才會套用「調整為 1 頁寬」
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins.top = 2.7 / 2.54
    ws.page_margins.bottom = 2.5 / 2.54
    ws.page_margins.left = 1.5 / 2.54
    ws.page_margins.right = 1.5 / 2.54
    ws.page_margins.header = 1.5 / 2.54
    ws.page_margins.footer = 1.0 / 2.54

    ws.oddHeader.center.text = f'&"Noto Sans TC,Bold"&16{report_title}'
    ws.oddFooter.left.text = '&"Noto Sans TC,Regular"&12中央健康保險署  政府資料開放平台 2026年資料'
    ws.oddFooter.right.text = '&"Noto Sans TC,Regular"&12https://data.gov.tw/dataset/22131\nhttps://data.gov.tw/dataset/23715'

    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    subtotal_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    total_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")

    # 字型與網頁預覽(FONT_FAMILY)、標題、頁尾統一使用 Noto Sans TC：
    # 微軟正黑體在沒有安裝該字型的系統（例如產生 PNG 預覽用的無頭瀏覽器／Excel 開啟環境）
    # 會對英文字母/數字這類非中文字元退回系統預設字型（如 Times New Roman），
    # 導致英文數字跟中文標題看起來像兩種字型；Noto Sans TC 對拉丁字母與數字也有完整設計，
    # 才能讓英文、數字跟標題、頁尾維持同一種字型風格。
    font_family = "Noto Sans TC"
    header_font = Font(name=font_family, size=12, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=12)
    bold_font = Font(name=font_family, size=12, bold=True)

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    thin_side = Side(border_style="thin", color="D9D9D9")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 標題列
    header_texts = list(row_fields)
    max_header_lines = 1
    for c in display_cols:
        lines = pretty_header(c)
        max_header_lines = max(max_header_lines, len(lines))
        header_texts.append("\n".join(lines))
    ws.append(header_texts)
    # 標題列高度依「最多行數的表頭」動態調整（例如含包裹占比/成長率欄位變成三行後，
    # 原本固定 60 會太矮導致文字被裁切），每多一行大約多需要 20 左右的高度。
    ws.row_dimensions[1].height = max(60, 20 * max_header_lines + 15)
    ws.print_title_rows = "1:1"

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = cell_border

    current_row = 2
    for r in rows:
        row_vals = []
        link_col_for_row = None  # (欄位索引, 原始網址)：許可證連結欄要在儲存格顯示「連結」並附上超連結
        for f in row_fields:
            # 空字串必須寫成 None，而不是 ""：openpyxl 對空字串會產生
            # <c t="inlineStr"></c>（宣告是字串型別卻沒有 <is> 內容），
            # 這是不合法的 OOXML，Excel 開啟時就會跳出「部分內容有問題」的修復提示，
            # 且修復過程中可能連帶把頁首/頁尾等設定一併清掉。寫 None 則完全不產生該儲存格，才是合法的空白儲存格。
            val = r["values"].get(f, "")
            if f == LINK_FIELD and val not in ("", None):
                link_col_for_row = (len(row_vals) + 1, val)
                row_vals.append(LINK_DISPLAY_TEXT)
            else:
                row_vals.append(val if val != "" else None)
        for c in display_cols:
            if c in extra_cols:
                v, ok = safe_numeric(r["sums"].get(c, 0))
                row_vals.append(v if ok else (str(v) if v != "" else None))
            else:
                raw = r["sums"].get(c, "")
                if raw == "":
                    row_vals.append(None)
                else:
                    v, ok = safe_numeric(raw)
                    row_vals.append(v if ok else str(v))
        ws.append(row_vals)
        ws.row_dimensions[current_row].height = 35

        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=i)
            if i <= len(row_fields):
                cell.alignment = left_align if i == 1 else center_align
                cell.font = bold_font if r["type"] != "data" else data_font
            elif h in extra_cols:
                cell.number_format = "0.0%"
                cell.alignment = right_align
                cell.font = bold_font if r["type"] != "data" else data_font
            else:
                cell.number_format = "#,##0"
                cell.alignment = right_align
                cell.font = bold_font if r["type"] != "data" else data_font
            cell.border = cell_border
            if r["type"] == "subtotal":
                cell.fill = subtotal_fill
            elif r["type"] == "total":
                cell.fill = total_fill

        if link_col_for_row:
            link_col_idx, link_url = link_col_for_row
            link_cell = ws.cell(row=current_row, column=link_col_idx)
            link_cell.hyperlink = link_url
            link_cell.font = Font(name=font_family, size=12, underline="single", color="1F497D")

        current_row += 1

    # 同類型欄位（數量／占比／成長率）欄寬要一致。
    # 舊版這裡用 headers[col_idx-1] 跟 "占比(%)"/"成長率(%)" 這種寫死的短字串比對，
    # 但 headers 存的其實是完整內部欄名（例如 "2025年占比(%)"），字串不可能完全相等，
    # 所以這個分支形同虛設──改用「這個欄名是否屬於 pct_cols / growth_cols」來正確判斷類別。
    for col_idx, col_key in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        if col_idx == 1:
            ws.column_dimensions[col_letter].width = 30
        elif col_idx <= len(row_fields):
            ws.column_dimensions[col_letter].width = 18
        elif col_key in growth_cols:
            ws.column_dimensions[col_letter].width = 16
        elif col_key in pct_cols:
            # 寬度需容得下最長的占比標題「2026年推估占比(%)」拆成兩行，
            # 原本 14 太窄會被 Excel 自動換成三行（"2026年"/"推估"/"占比(%)"），改成 18 才夠兩行顯示
            ws.column_dimensions[col_letter].width = 18
        else:
            ws.column_dimensions[col_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================
# Streamlit 主畫面（單一分析功能：廠商分析，不需選擇分析功能）
# ============================================================

st.title("📊 健保資料庫分析工具")
st.caption("申報量已更新至2026年6月。")

df_raw = load_data(DATA_FILE)

if df_raw.empty:
    st.warning(f"找不到資料檔案「{DATA_FILE}」，請確認檔案已放置於工作目錄。")
elif "成分" not in df_raw.columns:
    st.error(f"資料檔案缺少「成分」欄位，實際欄位為：{list(df_raw.columns)}")
else:
    import re

    def _parse_ingredient(part):
        # 嘗試把單一成分拆成「藥名」與「劑量數值」兩部分，例如
        # "Amlodipine 10mg" -> 藥名 "amlodipine"、劑量 10.0
        # 拆不出劑量數值時（格式特殊或無劑量），劑量視為 0，仍以藥名排序。
        part = part.strip()
        m = re.match(r"^(.*?)\s*([\d.]+)\s*\D*$", part)
        if m and m.group(1).strip():
            drug_name = m.group(1).strip().lower()
            try:
                dose_value = float(m.group(2))
            except ValueError:
                dose_value = 0.0
        else:
            drug_name = part.lower()
            dose_value = 0.0
        return drug_name, dose_value

    def _combo_sort_key(name):
        # 複方成分常以「+」分隔多個單方組成，原本排序是把整段（含劑量）當字串排序，
        # 導致「Amlodipine 10mg+Atorvastatin 10mg」跟「Amlodipine 5mg+Atorvastatin 10mg」
        # 因為劑量數字不同，被排到很遠的地方，即使是同一組複方也無法排在一起。
        #
        # 改成先只用「藥名」（忽略劑量、忽略原始順序）分組排序，讓同一組複方的所有劑量
        # 組合排在一起；同一組內再依各藥名對應的劑量數值排序，讓結果穩定、好尋找。
        parsed = sorted(
            (_parse_ingredient(p) for p in name.split("+")),
            key=lambda x: x[0],
        )
        drug_names = tuple(p[0] for p in parsed)
        doses = tuple(p[1] for p in parsed)
        return (drug_names, doses, name.lower())

    comp_options = sorted(
        [c for c in df_raw["成分"].dropna().unique() if c],
        key=_combo_sort_key,
    )

    if "comps_selected_persist" not in st.session_state:
        st.session_state["comps_selected_persist"] = []

    # 搜尋關鍵字用獨立的 text_input 元件保存，不會因為在下方勾選/取消勾選成分而被清空，
    # 這樣選了一項之後，下拉式清單仍會停留在同一組關鍵字的搜尋結果，不必重新輸入
    search_kw = st.text_input(
        "第1步：輸入關鍵字搜尋成分",
        key="comp_search_kw",
        placeholder="輸入關鍵字，例如 Ezetimibe",
    )

    if search_kw.strip():
        kw = search_kw.strip().lower()

        def _relevance_rank(name):
            n = name.lower()
            if n == kw:
                return 0
            if n.startswith(kw):
                return 1
            return 2

        filtered_options = [c for c in comp_options if kw in c.lower()]
        filtered_options.sort(key=lambda c: (_relevance_rank(c), _combo_sort_key(c)))
    else:
        filtered_options = []

    persisted_selected = [c for c in st.session_state["comps_selected_persist"] if c in comp_options]
    # 合併「搜尋結果」與「目前已勾選項目」並去重，確保已勾選成分不會因為搜尋字變動而從清單消失，
    # 也避免同一個成分同時出現在兩邊來源時被重複列出
    combined_options = list(dict.fromkeys(filtered_options + persisted_selected))

    comps_selected = st.multiselect(
        "第2步：勾選成分品項 (可多選)",
        options=combined_options,
        default=persisted_selected,
        key="comps_select_widget",
    )
    st.session_state["comps_selected_persist"] = comps_selected

    if not comps_selected:
        st.info("請先選擇至少一個成分，才會顯示後續的欄位設定。")
    else:
        df_comp = df_raw[df_raw["成分"].isin(comps_selected)]

        st.markdown("### 🧩 第2步：拖曳欄位到「報表欄位」，並排序")
        st.caption("僅開放「成分、劑型、劑量、規格、廠商、健保代碼、商品名、健保價、ATC碼、許可證連結」十個欄位可拖曳排列。")

        dnd_key = f"dnd_{'_'.join(sorted(comps_selected))}"
        state_key = f"pivot_state_{dnd_key}"
        DEFAULT_SELECTED_FIELDS = ["成分", "劑型", "劑量", "廠商"]
        if state_key not in st.session_state:
            default_selected = [c for c in DEFAULT_SELECTED_FIELDS if c in FIXED_FIELDS]
            default_available = [c for c in FIXED_FIELDS if c not in default_selected]
            st.session_state[state_key] = {"available": default_available, "selected": default_selected}
        else:
            prev = st.session_state[state_key]
            avail = [c for c in prev["available"] if c in FIXED_FIELDS]
            sel = [c for c in prev["selected"] if c in FIXED_FIELDS]
            known = set(avail) | set(sel)
            avail += [c for c in FIXED_FIELDS if c not in known]
            st.session_state[state_key] = {"available": avail, "selected": sel}

        use_fallback = st.checkbox(
            "⚠️ 拖曳排序出現錯誤時，改用勾選方式選擇欄位 (勾選順序＝欄位順序)",
            value=False, key=f"use_fallback_{dnd_key}",
        )

        if HAS_SORTABLES and not use_fallback:
            containers = sort_items(
                [
                    {"header": "📋 可用欄位 (拖曳到下方使用)", "items": st.session_state[state_key]["available"]},
                    {"header": "📊 報表欄位 (由左到右排列)", "items": st.session_state[state_key]["selected"]},
                ],
                multi_containers=True, direction="horizontal", key=dnd_key,
            )
            if containers and len(containers) > 1:
                st.session_state[state_key] = {"available": containers[0]["items"], "selected": containers[1]["items"]}
            row_fields = st.session_state[state_key]["selected"]
        elif HAS_SORTABLES:
            sel = st.multiselect(
                "選擇報表欄位 (依勾選順序排列)",
                options=FIXED_FIELDS, default=st.session_state[state_key]["selected"],
                key=f"fallback_fields_{dnd_key}",
            )
            st.session_state[state_key] = {"available": [c for c in FIXED_FIELDS if c not in sel], "selected": sel}
            row_fields = sel
        else:
            st.error("尚未安裝 streamlit-sortables 套件，暫以勾選方式呈現，請於 requirements.txt 加入 streamlit-sortables 後重新部署即可拖曳。")
            row_fields = st.multiselect("選擇報表欄位", options=FIXED_FIELDS, key="fallback_fields_static")

        if not row_fields:
            st.info("請至少拖曳一個欄位到「報表欄位」區塊。")
        else:
            st.markdown("### 🔍 第3步：逐欄篩選與合計 (點欄位下方的「🔽 篩選 / 合計」展開設定)")
            st.caption("💡 篩選會彼此連動：某欄位選擇後，其他欄位只會列出仍有對應資料的選項。")

            current_filters = {}
            for f in row_fields:
                key = f"filt_{dnd_key}_{f}"
                if key in st.session_state and st.session_state[key]:
                    current_filters[f] = st.session_state[key]

            filter_ui_cols = st.columns(len(row_fields))
            filters = {}
            subtotal_fields_selected = []
            for i, f in enumerate(row_fields):
                with filter_ui_cols[i]:
                    st.markdown(f"**{f}**")
                    with st.expander("🔽 篩選 / 合計", expanded=False):
                        df_scope = df_comp
                        for other_col, other_sel in current_filters.items():
                            if other_col != f and other_sel:
                                df_scope = df_scope[df_scope[other_col].isin(other_sel)]
                        options = sorted(
                            [v for v in df_scope[f].dropna().unique() if str(v).strip() != ""],
                            key=parse_dosage if f == "劑量" else None,
                        )

                        key = f"filt_{dnd_key}_{f}"
                        if key in st.session_state:
                            st.session_state[key] = [v for v in st.session_state[key] if v in options]

                        sel = st.multiselect(f"篩選「{f}」(不選代表全選)", options=options, key=key)
                        if sel:
                            filters[f] = sel
                        if st.checkbox("Σ 此欄要合計", key=f"sub_{dnd_key}_{f}"):
                            subtotal_fields_selected.append(f)
            subtotal_fields = [f for f in row_fields if f in subtotal_fields_selected]

            df_filtered = df_comp.copy()
            for col, sel in filters.items():
                df_filtered = df_filtered[df_filtered[col].isin(sel)]

            st.caption(f"篩選後共 **{len(df_filtered):,}** 筆資料")

            st.markdown("### 🧮 第4步：選擇要加總的數值欄位 (預設帶入 2024~2026年推估 含包裹數量)")
            st.markdown(
                f"<div style='background-color:#EAF1FB;border-left:4px solid #1F497D;"
                f"padding:8px 12px;border-radius:4px;font-size:13px;color:#333333;'>{QTY_DEFINITION_NOTE}</div>",
                unsafe_allow_html=True,
            )

            # 這裡顯示的欄位名稱要跟報表結果的欄位標題一致（例如「2026年(1-6月) 數量」），
            # 而不是內部使用的原始欄名（例如「2026年申報量」），避免使用者選欄位時對不上結果
            def qty_display_label(internal_col):
                year, label = QTY_DISPLAY_LABEL.get(internal_col, (None, internal_col))
                return f"{year} {label}" if year else label

            qty_options = list(QTY_COL_MAP.values()) + [EST_QTY_COL, EST_BUNDLE_QTY_COL]
            qty_label_map = {c: qty_display_label(c) for c in qty_options}
            qty_label_to_internal = {v: k for k, v in qty_label_map.items()}

            # 跟第2步的拖曳介面呈現方式一致：未選用的欄位放在上方「可用欄位」，
            # 已選用、會加總／顯示的欄位放在下方，且可拖曳調整加總／顯示順序。
            # 預設帶入「2024、2025、2026年推估」三個含包裹數量欄位，其餘欄位（含一般醫令量、
            # 2023年含包裹、2026年(1-6月)含包裹）預設放進「可用欄位」，使用者需要時自行拖曳加入。
            DEFAULT_SELECTED_QTY = ["2024年含包裹申報量", "2025年含包裹申報量", EST_BUNDLE_QTY_COL]
            qty_state_key = f"qty_state_{dnd_key}"
            if qty_state_key not in st.session_state:
                default_selected = [c for c in DEFAULT_SELECTED_QTY if c in qty_options]
                default_available = [c for c in qty_options if c not in DEFAULT_SELECTED_QTY]
                st.session_state[qty_state_key] = {"available": default_available, "selected": default_selected}
            else:
                prev_qty = st.session_state[qty_state_key]
                avail_q = [c for c in prev_qty["available"] if c in qty_options]
                sel_q = [c for c in prev_qty["selected"] if c in qty_options]
                known_q = set(avail_q) | set(sel_q)
                avail_q += [c for c in qty_options if c not in known_q]
                st.session_state[qty_state_key] = {"available": avail_q, "selected": sel_q}

            use_qty_fallback = st.checkbox(
                "⚠️ 拖曳排序出現錯誤時，改用勾選方式選擇欄位 (勾選順序＝欄位順序)",
                value=False, key=f"use_qty_fallback_{dnd_key}",
            )

            if HAS_SORTABLES and not use_qty_fallback:
                qty_containers = sort_items(
                    [
                        {
                            "header": "📋 可用欄位 (拖曳到下方使用)",
                            "items": [qty_label_map[c] for c in st.session_state[qty_state_key]["available"]],
                        },
                        {
                            "header": "🧮 要加總的數值欄位 (由左到右排列)",
                            "items": [qty_label_map[c] for c in st.session_state[qty_state_key]["selected"]],
                        },
                    ],
                    multi_containers=True, direction="horizontal", key=f"qty_dnd_{dnd_key}",
                )
                if qty_containers and len(qty_containers) > 1:
                    avail_internal = [qty_label_to_internal.get(lbl, lbl) for lbl in qty_containers[0]["items"]]
                    sel_internal = [qty_label_to_internal.get(lbl, lbl) for lbl in qty_containers[1]["items"]]
                    st.session_state[qty_state_key] = {"available": avail_internal, "selected": sel_internal}
                value_cols = st.session_state[qty_state_key]["selected"]
                st.caption("💡 可直接拖曳上方欄位調整加總／顯示順序。")
            elif HAS_SORTABLES:
                sel_v = st.multiselect(
                    "選擇要加總的數值欄位 (依勾選順序排列)",
                    options=qty_options, default=st.session_state[qty_state_key]["selected"],
                    format_func=lambda c: qty_label_map.get(c, c), key=f"qty_fallback_{dnd_key}",
                )
                st.session_state[qty_state_key] = {
                    "available": [c for c in qty_options if c not in sel_v], "selected": sel_v,
                }
                value_cols = sel_v
            else:
                st.error("尚未安裝 streamlit-sortables 套件，暫以勾選方式呈現，請於 requirements.txt 加入 streamlit-sortables 後重新部署即可拖曳。")
                value_cols = st.multiselect(
                    "選擇要加總的數值欄位", options=qty_options, default=qty_options,
                    format_func=lambda c: qty_label_map.get(c, c), key=f"qty_fallback_static_{dnd_key}",
                )

            if not value_cols:
                st.info("請至少選擇一個要加總的數值欄位。")

            # 年度占比／年度成長率選單完全獨立於上方「要加總的數值欄位」：不論使用者有沒有勾選要
            # 「顯示」某年度（或某年度含包裹）的數量欄位，只要在下面選了該年度，就會用該年度
            # （該類型）的原始資料正確算出占比／成長率，不會發生分母抓不到資料而變成 0% 的情況。
            has_qty = True  # 年度選項固定為 2023~2026，恆為可選；停用邏輯改成看有無勾選成長率本身即可

            pct_years = st.multiselect(
                "➕ 加入年度占比(%) (可只選需要的年份；含包裹／不含包裹分開列，可各自單選或都選)",
                options=PCT_YEAR_CODE_OPTIONS,
                default=["2025含包裹", EST_BUNDLE_PCT_YEAR],
                key=f"pct_{dnd_key}",
                format_func=pct_year_code_label,
            )
            add_growth = st.checkbox("➕ 加入年度成長率(%)", value=False, key=f"growth_{dnd_key}")

            # 成長率只提供「醫令量→醫令量」或「含包裹醫令量→含包裹醫令量」這種同類型的年度區間，
            # 不會出現「醫令量→含包裹醫令量」這種跨類型比較（詳見 GROWTH_PAIR_CODE_OPTIONS 的定義）。
            growth_pairs = []
            if add_growth:
                growth_pairs = st.multiselect(
                    "選擇要顯示的成長率年度區間 (可複選；僅提供同類型「醫令量↔醫令量」"
                    "或「含包裹↔含包裹」的區間，不提供跨類型比較)",
                    options=GROWTH_PAIR_CODE_OPTIONS,
                    default=[EST_BUNDLE_GROWTH_PAIR],
                    format_func=lambda p: growth_pair_code_label(*p),
                    key=f"growth_pairs_{dnd_key}",
                )

            # 標題與檔名都要反映「篩選了哪些欄位」：成分一定顯示；其餘欄位（劑型/劑量/規格/廠商）
            # 只有在使用者有實際篩選時才附加上去，沒特別篩選的欄位就不出現在標題/檔名裡
            filter_desc_parts_for_filename = []
            filter_desc_parts_for_title = []
            for col in row_fields:
                if col == "成分":
                    continue
                if col in filters and filters[col]:
                    filter_desc_parts_for_filename.append("_".join(filters[col]))
                    filter_desc_parts_for_title.append("、".join(filters[col]))

            filename_parts = list(comps_selected) + filter_desc_parts_for_filename + ["廠商申報量排名"]
            report_title = "、".join(comps_selected) + "".join(filter_desc_parts_for_title) + "廠商申報量排名"

            MAX_FILENAME_BYTES = 200  # 留安全餘裕，避免超過檔案系統上限 (通常 255 bytes)，
            # 也避免選太多成分/篩選條件時檔名過長導致下載失敗
            ext_reserve = 5  # 保留給 ".xlsx" / ".png" 副檔名的空間

            def make_safe_filename(raw: str) -> str:
                """把任意字串（自動產生的標題，或使用者自訂的標題）轉成安全檔名：
                去除不合法字元，並在超過長度上限時截斷加上「...」。"""
                raw = re.sub(r'[\\/*?:"<>|]', "_", raw)
                if len(raw.encode("utf-8")) + ext_reserve <= MAX_FILENAME_BYTES:
                    return raw
                budget = MAX_FILENAME_BYTES - ext_reserve - 3  # 預留「...」空間
                truncated = raw
                while len(truncated.encode("utf-8")) > budget and len(truncated) > 0:
                    truncated = truncated[:-1]
                return truncated + "..."

            raw_filename = "_".join(filename_parts)
            safe_filename = make_safe_filename(raw_filename)

            st.divider()

            # 自訂標題／檔名：預設關閉，勾選後可自行輸入。一旦有輸入內容，
            # 預覽標題、Excel 標題／檔名、PNG 圖片檔名都會改套用這個自訂標題，
            # 不再使用依「成分＋篩選條件」自動組出來的標題／檔名。
            # 這個功能容易被忽略，所以加上底色的提示區塊讓使用者比較容易注意到有這個選項。
            st.markdown(
                "<div style='background-color:#FFF6DA;border-left:4px solid #E0A800;"
                "padding:8px 12px;border-radius:4px;font-size:13px;color:#333333;margin-bottom:6px;'>"
                "💡 如果想自行輸入報表標題／檔名，可以打開下方選項。</div>",
                unsafe_allow_html=True,
            )
            custom_title_enabled = st.checkbox(
                "✏️ 自訂標題／檔名 (勾選後可自行輸入，將套用於預覽標題、Excel、圖片檔名)",
                value=False, key=f"custom_title_toggle_{dnd_key}",
            )
            if custom_title_enabled:
                custom_title_input = st.text_input(
                    "輸入自訂標題／檔名",
                    value=report_title, key=f"custom_title_{dnd_key}",
                )
                if custom_title_input.strip():
                    report_title = custom_title_input.strip()
                    safe_filename = make_safe_filename(report_title)

            if value_cols and not df_filtered.empty:
                rows, pct_cols, growth_cols = build_nested_rows(
                    df_filtered, row_fields, subtotal_fields, value_cols, pct_years, growth_pairs,
                )
                st.markdown("### 📄 報表即時預覽")
                table_html = build_html_table(rows, row_fields, value_cols, pct_cols, growth_cols, report_title)
                st.markdown(table_html, unsafe_allow_html=True)

                st.markdown("### 📥 下載")
                dl_col1, dl_col2 = st.columns(2)
                safe_dnd_key = re.sub(r"[^0-9A-Za-z_]", "_", dnd_key)
                with dl_col1:
                    excel_bytes = generate_excel_bytes(rows, row_fields, value_cols, pct_cols, growth_cols, report_title)
                    render_excel_share(excel_bytes, safe_filename, uid=f"pivot_{safe_dnd_key}")
                with dl_col2:
                    components.html(
                        CAPTURE_HTML_TEMPLATE.format(table_html=table_html, filename=safe_filename, uid=f"pivot_{safe_dnd_key}"),
                        height=90,
                    )
            elif not value_cols:
                st.warning("⚠️ 請至少選擇一個要加總的數值欄位。")
            else:
                st.warning("❌ 篩選後無資料，請放寬篩選條件。")
